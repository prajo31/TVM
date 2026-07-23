"""
Time Value of Money (TVM) Calculator — for student learning
------------------------------------------------------------
Core concepts: Single Sum, Annuity (ordinary/due/growing), Perpetuity
(level/growing), APR <-> EAR conversion.

Real-world case studies: Mortgage/Loan amortization, Retirement
Planner (with inflation & salary-growth assumptions), Lottery lump
sum vs annuity, Credit Card APR vs EAR.

Every calculator shows the formula used, a plain-language
interpretation, and a cash-flow visual. The case-study tabs also
offer Excel/PDF downloads of their schedules.

Run with:  streamlit run tvm_calculator_app.py
"""

import io
import math

import pandas as pd
import streamlit as st

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter as PAGE_SIZE
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

st.set_page_config(page_title="Time Value of Money Calculator", page_icon="⏳", layout="wide")

# ============================================================
# FORMATTING HELPERS
# ============================================================

def fmt_money(x, decimals=2):
    """Formats currency for markdown-rendered text (st.success/warning/info/caption/write).
    Escapes the $ so Streamlit doesn't mistake it for LaTeX math delimiters — a sentence
    with two or more dollar amounts would otherwise get silently mangled."""
    if x is None:
        return "N/A"
    return f"\\${x:,.{decimals}f}"


def fmt_money_plain(x, decimals=2):
    """Unescaped currency for contexts that DON'T parse markdown: st.metric values,
    and text written into Excel cells or PDF tables."""
    if x is None:
        return "N/A"
    return f"${x:,.{decimals}f}"


def fmt_pct(x, decimals=2):
    if x is None:
        return "N/A"
    return f"{x * 100:,.{decimals}f}%"


def fmt_num(x, decimals=2):
    if x is None:
        return "N/A"
    return f"{x:,.{decimals}f}"


# ============================================================
# CORE TVM MATH ENGINE (pure functions — unit-tested independently)
# ============================================================

def fv_single(pv, i, n):
    return pv * (1 + i) ** n


def pv_single(fv, i, n):
    return fv / (1 + i) ** n


def n_single(pv, fv, i):
    if pv <= 0 or fv <= 0 or i <= -1 or i == 0:
        return None
    return math.log(fv / pv) / math.log(1 + i)


def rate_single(pv, fv, n):
    if pv <= 0 or fv <= 0 or n <= 0:
        return None
    return (fv / pv) ** (1 / n) - 1


def fv_annuity(pmt, i, n, due=False, g=0.0):
    if n <= 0:
        return 0.0
    if g == 0:
        fv = pmt * n if i == 0 else pmt * (((1 + i) ** n - 1) / i)
    else:
        if abs(i - g) < 1e-12:
            fv = pmt * n * (1 + i) ** (n - 1)
        else:
            fv = pmt * (((1 + i) ** n - (1 + g) ** n) / (i - g))
    if due:
        fv *= (1 + i)
    return fv


def pv_annuity(pmt, i, n, due=False, g=0.0):
    if n <= 0:
        return 0.0
    if g == 0:
        pv = pmt * n if i == 0 else pmt * (1 - (1 + i) ** -n) / i
    else:
        if abs(i - g) < 1e-12:
            pv = pmt * n / (1 + i)
        else:
            pv = (pmt / (i - g)) * (1 - ((1 + g) / (1 + i)) ** n)
    if due:
        pv *= (1 + i)
    return pv


def pmt_from_pv(pv, i, n, due=False):
    if n <= 0:
        return None
    pmt = pv / n if i == 0 else pv * i / (1 - (1 + i) ** -n)
    if due:
        pmt /= (1 + i)
    return pmt


def pmt_from_fv(fv, i, n, due=False):
    if n <= 0:
        return None
    pmt = fv / n if i == 0 else fv * i / ((1 + i) ** n - 1)
    if due:
        pmt /= (1 + i)
    return pmt


def pmt1_from_growing_fv(fv, i, n, g, due=False):
    if n <= 0:
        return None
    denom = n * (1 + i) ** (n - 1) if abs(i - g) < 1e-12 else ((1 + i) ** n - (1 + g) ** n) / (i - g)
    if denom == 0:
        return None
    pmt1 = fv / denom
    if due:
        pmt1 /= (1 + i)
    return pmt1


def _bisect(f, target, lo, hi, tol=1e-9, max_iter=200):
    flo, fhi = f(lo) - target, f(hi) - target
    if flo == 0:
        return lo
    if fhi == 0:
        return hi
    if (flo > 0) == (fhi > 0):
        return None
    for _ in range(max_iter):
        mid = (lo + hi) / 2
        fmid = f(mid) - target
        if abs(fmid) < tol:
            return mid
        if (fmid > 0) == (flo > 0):
            lo, flo = mid, fmid
        else:
            hi, fhi = mid, fmid
    return (lo + hi) / 2


def n_annuity(pmt, i, target_pv=None, target_fv=None, due=False, g=0.0):
    if g == 0 and target_fv is not None and i != 0:
        pmt_eff = pmt * (1 + i) if due else pmt
        inside = 1 + target_fv * i / pmt_eff
        return math.log(inside) / math.log(1 + i) if inside > 0 else None
    if g == 0 and target_pv is not None and i != 0:
        pmt_eff = pmt * (1 + i) if due else pmt
        inside = 1 - target_pv * i / pmt_eff
        return -math.log(inside) / math.log(1 + i) if inside > 0 else None
    if target_fv is not None:
        return _bisect(lambda n: fv_annuity(pmt, i, n, due=due, g=g), target_fv, 1e-6, 1000)
    if target_pv is not None:
        return _bisect(lambda n: pv_annuity(pmt, i, n, due=due, g=g), target_pv, 1e-6, 1000)
    return None


def rate_annuity(pmt, n, target_pv=None, target_fv=None, due=False, g=0.0):
    lo, hi = max(-0.99, g - 0.99), 10.0
    if target_pv is not None:
        return _bisect(lambda i: pv_annuity(pmt, i, n, due=due, g=g), target_pv, lo, hi)
    if target_fv is not None:
        return _bisect(lambda i: fv_annuity(pmt, i, n, due=due, g=g), target_fv, lo, hi)
    return None


def pv_perpetuity(pmt, i, g=0.0):
    return None if i <= g else pmt / (i - g)


def pmt_from_perpetuity(pv, i, g=0.0):
    return None if i <= g else pv * (i - g)


def rate_from_perpetuity(pv, pmt, g=0.0):
    return None if pv <= 0 else pmt / pv + g


def ear_from_apr(apr, m):
    return math.exp(apr) - 1 if m is None else (1 + apr / m) ** m - 1


def apr_from_ear(ear, m):
    return math.log(1 + ear) if m is None else m * ((1 + ear) ** (1 / m) - 1)


def fisher_real_rate(nominal_rate, inflation_rate):
    return (1 + nominal_rate) / (1 + inflation_rate) - 1


def rule_of_72(rate_pct):
    return None if rate_pct <= 0 else 72 / rate_pct


def amortization_schedule(principal, annual_rate, years, payments_per_year=12, extra_payment=0.0):
    n = int(round(years * payments_per_year))
    i = annual_rate / payments_per_year
    base_pmt = pmt_from_pv(principal, i, n)
    schedule = []
    balance = principal
    total_interest = 0.0
    period = 0
    while balance > 0.01 and period < n * 3:
        period += 1
        interest = balance * i
        principal_pmt = base_pmt - interest + extra_payment
        if principal_pmt > balance:
            principal_pmt = balance
        payment = interest + principal_pmt
        balance -= principal_pmt
        total_interest += interest
        schedule.append(dict(period=period, payment=payment, interest=interest,
                              principal=principal_pmt, balance=max(balance, 0.0)))
        if balance <= 0.01:
            break
    summary = dict(scheduled_payment=base_pmt, n_scheduled=n, n_actual=len(schedule),
                   total_interest=total_interest, total_paid=total_interest + principal,
                   payoff_years=len(schedule) / payments_per_year)
    return schedule, summary


def growing_annuity_projection(pv0, pmt1, i, g, n):
    rows = []
    balance = pv0
    contribution = pmt1
    for year in range(1, n + 1):
        balance = balance * (1 + i) + contribution
        rows.append(dict(year=year, contribution=contribution, balance=balance))
        contribution *= (1 + g)
    return rows


def lottery_pv_of_annuity(first_payment, n_payments, growth_rate, discount_rate):
    return pv_annuity(first_payment, discount_rate, n_payments, due=True, g=growth_rate)


def lottery_breakeven_rate(first_payment, n_payments, growth_rate, lump_sum):
    return _bisect(lambda i: pv_annuity(first_payment, i, n_payments, due=True, g=growth_rate),
                    lump_sum, 1e-9, 5.0)


# ============================================================
# EXCEL / PDF EXPORT HELPERS
# ============================================================

def _style_header_row(ws, row, headers, n_cols):
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def build_amortization_excel(title, summary_rows, schedule):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    title_font = Font(name="Arial", size=14, bold=True)
    normal_font = Font(name="Arial", size=10)
    ws.merge_cells("A1:B1")
    ws["A1"] = title
    ws["A1"].font = title_font
    r = 3
    for label, value in summary_rows:
        ws.cell(row=r, column=1, value=label).font = normal_font
        ws.cell(row=r, column=2, value=value).font = normal_font
        r += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20

    ws2 = wb.create_sheet("Amortization Schedule")
    headers = ["Period", "Payment", "Interest", "Principal", "Remaining Balance"]
    _style_header_row(ws2, 1, headers, len(headers))
    normal_font2 = Font(name="Arial", size=10)
    for r_idx, row in enumerate(schedule, start=2):
        ws2.cell(row=r_idx, column=1, value=row["period"]).font = normal_font2
        ws2.cell(row=r_idx, column=2, value=round(row["payment"], 2)).font = normal_font2
        ws2.cell(row=r_idx, column=3, value=round(row["interest"], 2)).font = normal_font2
        ws2.cell(row=r_idx, column=4, value=round(row["principal"], 2)).font = normal_font2
        ws2.cell(row=r_idx, column=5, value=round(row["balance"], 2)).font = normal_font2
    for i, w in enumerate([10, 14, 14, 14, 16], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_simple_pdf(title, subtitle, summary_rows, table_headers=None, table_rows=None, footnote=None):
    """Generic PDF builder: title + key/value summary + optional data table."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=PAGE_SIZE, leftMargin=40, rightMargin=40, topMargin=48, bottomMargin=36)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("T", parent=styles["Title"], fontSize=16, spaceAfter=4)
    sub_style = ParagraphStyle("S", parent=styles["Normal"], fontSize=9, textColor=colors.grey, spaceAfter=14)
    section_style = ParagraphStyle("Sec", parent=styles["Heading2"], fontSize=12, spaceBefore=10, spaceAfter=6,
                                    textColor=colors.HexColor("#1F4E78"))
    footnote_style = ParagraphStyle("F", parent=styles["Normal"], fontSize=7.5, textColor=colors.grey)

    story = [Paragraph(title, title_style)]
    if subtitle:
        story.append(Paragraph(subtitle, sub_style))

    if summary_rows:
        data = [["Item", "Value"]] + [[k, v] for k, v in summary_rows]
        t = Table(data, colWidths=[220, 200])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F6FA")]),
            ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(t)
        story.append(Spacer(1, 12))

    if table_headers and table_rows:
        story.append(Paragraph("Detailed Schedule", section_style))
        data2 = [table_headers] + table_rows
        col_w = [500 / len(table_headers)] * len(table_headers)
        t2 = Table(data2, colWidths=col_w, repeatRows=1)
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F6FA")]),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(t2)

    if footnote:
        story.append(Spacer(1, 10))
        story.append(Paragraph(footnote, footnote_style))

    doc.build(story)
    buf.seek(0)
    return buf


# ============================================================
# SIDEBAR
# ============================================================
with st.sidebar:
    st.header("📘 About this tool")
    st.markdown(
        """
        A Time-Value-of-Money calculator built for learning:

        **Core concepts**
        - Single Sum (lump sum PV/FV)
        - Annuity (ordinary, due, growing)
        - Perpetuity (level, growing)
        - APR ↔ EAR conversion

        **Real-world case studies**
        - Mortgage / loan amortization
        - Retirement planner (with inflation & salary growth)
        - Lottery: lump sum vs. annuity
        - Credit card: APR vs. EAR

        Every result includes the formula used and a plain-language
        interpretation — not just a number.
        """
    )
    st.divider()
    st.subheader("⚡ Quick tools")
    r72 = st.number_input("Rule of 72: rate (%)", value=6.0, step=0.5, key="sidebar_r72")
    years72 = rule_of_72(r72)
    if years72:
        exact = n_single(1, 2, r72 / 100)
        st.caption(f"≈ **{years72:.1f} years** to double (exact: {exact:.1f} years)")
    st.caption("Ratio thresholds and case-study defaults are illustrative, editable examples — not financial advice.")

st.title("⏳ Time Value of Money Calculator")
st.caption("Solve core TVM problems, then apply them to real-world mortgage, retirement, lottery, and credit card decisions.")

tab_single, tab_annuity, tab_perp, tab_aprear, tab_mortgage, tab_retire, tab_lottery, tab_cc = st.tabs(
    ["🧮 Single Sum", "📈 Annuity", "♾️ Perpetuity", "🔄 APR ↔ EAR",
     "🏠 Mortgage/Loan", "🎯 Retirement", "🎰 Lottery", "💳 Credit Card"]
)

# ============================================================
# TAB: SINGLE SUM
# ============================================================
with tab_single:
    st.subheader("🧮 Single Sum (Lump Sum)")
    st.write("Given any three of Present Value, Future Value, rate per period, and number of periods, solve for the fourth.")

    solve_for = st.selectbox("Solve for:", ["Future Value (FV)", "Present Value (PV)",
                                             "Rate per Period (I/Y)", "Number of Periods (N)"], key="ss_solve")
    c1, c2 = st.columns(2)
    pv = fv = rate_pct = n = None
    with c1:
        if solve_for != "Present Value (PV)":
            pv = st.number_input("Present Value (PV)", value=1000.0, step=100.0, key="ss_pv")
        if solve_for != "Rate per Period (I/Y)":
            rate_pct = st.number_input("Rate per Period, I/Y (%)", value=5.0, step=0.25, key="ss_rate",
                                        help="For a monthly problem, this is the MONTHLY rate (annual ÷ 12).")
    with c2:
        if solve_for != "Number of Periods (N)":
            n = st.number_input("Number of Periods (N)", value=10.0, step=1.0, min_value=0.0, key="ss_n")
        if solve_for != "Future Value (FV)":
            fv = st.number_input("Future Value (FV)", value=1628.89, step=100.0, key="ss_fv")

    result, err = None, None
    if solve_for == "Future Value (FV)":
        result = fv_single(pv, rate_pct / 100, n)
    elif solve_for == "Present Value (PV)":
        result = pv_single(fv, rate_pct / 100, n)
    elif solve_for == "Rate per Period (I/Y)":
        if pv <= 0 or fv <= 0 or n <= 0:
            err = "PV, FV, and N must all be positive to solve for the rate."
        else:
            result = rate_single(pv, fv, n)
    else:
        if rate_pct == 0:
            err = "Rate can't be 0% when solving for N."
        elif pv <= 0 or fv <= 0:
            err = "PV and FV must be positive to solve for N."
        else:
            result = n_single(pv, fv, rate_pct / 100)

    st.divider()
    if err:
        st.error(err)
    elif result is None:
        st.error("Couldn't compute a result with these inputs — try adjusting the values.")
    else:
        flows = {}
        if solve_for == "Future Value (FV)":
            st.metric("Future Value (FV)", fmt_money_plain(result))
            st.success(f"Investing {fmt_money(pv)} today at {rate_pct:.2f}% per period for {n:.0f} periods "
                       f"grows to **{fmt_money(result)}** — a gain of {fmt_money(result - pv)} "
                       f"({(result / pv - 1) * 100:.1f}% total return).")
            flows = {0: -pv, int(round(n)): result}
            formula = "FV = PV × (1 + i)ⁿ"
            sub = f"FV = {fmt_money(pv)} × (1 + {rate_pct/100:.4f})^{n:.0f} = {fmt_money(result)}"
        elif solve_for == "Present Value (PV)":
            st.metric("Present Value (PV)", fmt_money_plain(result))
            st.success(f"To have {fmt_money(fv)} in {n:.0f} periods at {rate_pct:.2f}% per period, "
                       f"you'd need to invest **{fmt_money(result)}** today.")
            flows = {0: -result, int(round(n)): fv}
            formula = "PV = FV ÷ (1 + i)ⁿ"
            sub = f"PV = {fmt_money(fv)} ÷ (1 + {rate_pct/100:.4f})^{n:.0f} = {fmt_money(result)}"
        elif solve_for == "Rate per Period (I/Y)":
            st.metric("Rate per Period (I/Y)", fmt_pct(result))
            st.success(f"Growing {fmt_money(pv)} into {fmt_money(fv)} over {n:.0f} periods requires "
                       f"a periodic rate of **{fmt_pct(result)}**.")
            flows = {0: -pv, int(round(n)): fv}
            formula = "i = (FV ÷ PV)^(1/n) − 1"
            sub = f"i = ({fmt_money(fv)} ÷ {fmt_money(pv)})^(1/{n:.0f}) − 1 = {fmt_pct(result)}"
        else:
            st.metric("Number of Periods (N)", f"{result:.2f}")
            st.success(f"At {rate_pct:.2f}% per period, it takes about **{result:.2f} periods** "
                       f"to grow {fmt_money(pv)} into {fmt_money(fv)}.")
            flows = {0: -pv, int(round(result)): fv}
            formula = "n = ln(FV ÷ PV) ÷ ln(1 + i)"
            sub = f"n = ln({fmt_money(fv)} ÷ {fmt_money(pv)}) ÷ ln(1 + {rate_pct/100:.4f}) = {result:.2f}"

        with st.expander("📐 Formula used"):
            st.code(formula)
            st.caption(sub)

        st.caption("Cash-flow timeline:")
        chart_df = pd.DataFrame({"Cash Flow": flows}).sort_index()
        st.bar_chart(chart_df)

# ============================================================
# TAB: ANNUITY
# ============================================================
with tab_annuity:
    st.subheader("📈 Annuity")
    st.write("A series of equal (or steadily growing) payments at regular intervals.")

    ct1, ct2 = st.columns(2)
    with ct1:
        timing = st.radio("Payment timing", ["Ordinary (end of period)", "Due (start of period)"],
                           key="an_timing", horizontal=True)
    with ct2:
        is_growing = st.checkbox("Payments grow each period (growing annuity)", key="an_growing")
    due = timing.startswith("Due")

    solve_for = st.selectbox("Solve for:", ["Future Value (FV)", "Present Value (PV)", "Payment (PMT)",
                                             "Number of Periods (N)", "Rate per Period (I/Y)"], key="an_solve")

    reference = None
    if solve_for in ("Payment (PMT)", "Number of Periods (N)", "Rate per Period (I/Y)"):
        reference = st.radio("Your known amount is a:", ["Present Value (PV)", "Future Value (FV)"],
                              key="an_ref", horizontal=True)

    c1, c2, c3 = st.columns(3)
    pmt = rate_pct = n = pv_target = fv_target = g_pct = None
    with c1:
        if solve_for != "Payment (PMT)":
            pmt = st.number_input("Payment per period (PMT)" + (" — first payment" if is_growing else ""),
                                   value=100.0, step=10.0, key="an_pmt")
        if is_growing:
            g_pct = st.number_input("Growth Rate per Period (%)", value=2.0, step=0.25, key="an_g")
        else:
            g_pct = 0.0
    with c2:
        if solve_for != "Number of Periods (N)":
            n = st.number_input("Number of Periods (N)", value=10.0, step=1.0, min_value=0.0, key="an_n")
        if solve_for != "Rate per Period (I/Y)":
            rate_pct = st.number_input("Rate per Period, I/Y (%)", value=5.0, step=0.25, key="an_rate")
    with c3:
        need_pv = (solve_for == "Present Value (PV)") or (reference == "Present Value (PV)")
        need_fv = (solve_for == "Future Value (FV)") or (reference == "Future Value (FV)")
        if solve_for == "Present Value (PV)" or reference == "Present Value (PV)":
            if solve_for != "Present Value (PV)":
                pv_target = st.number_input("Present Value (PV) — known amount", value=1000.0, step=100.0, key="an_pv")
        if solve_for == "Future Value (FV)" or reference == "Future Value (FV)":
            if solve_for != "Future Value (FV)":
                fv_target = st.number_input("Future Value (FV) — known amount", value=1000.0, step=100.0, key="an_fv")

    result, err = None, None
    g = (g_pct or 0.0) / 100
    try:
        if solve_for == "Future Value (FV)":
            result = fv_annuity(pmt, rate_pct / 100, n, due=due, g=g)
        elif solve_for == "Present Value (PV)":
            result = pv_annuity(pmt, rate_pct / 100, n, due=due, g=g)
        elif solve_for == "Payment (PMT)":
            if reference == "Present Value (PV)":
                if g == 0:
                    result = pmt_from_pv(pv_target, rate_pct / 100, n, due=due)
                else:
                    err = "Solving PMT from a PV target isn't supported for growing annuities in this tool — try Future Value instead, or switch off growth."
            else:
                if g == 0:
                    result = pmt_from_fv(fv_target, rate_pct / 100, n, due=due)
                else:
                    result = pmt1_from_growing_fv(fv_target, rate_pct / 100, n, g, due=due)
        elif solve_for == "Number of Periods (N)":
            if reference == "Present Value (PV)":
                result = n_annuity(pmt, rate_pct / 100, target_pv=pv_target, due=due, g=g)
            else:
                result = n_annuity(pmt, rate_pct / 100, target_fv=fv_target, due=due, g=g)
        else:  # Rate per Period
            if reference == "Present Value (PV)":
                result = rate_annuity(pmt, n, target_pv=pv_target, due=due, g=g)
            else:
                result = rate_annuity(pmt, n, target_fv=fv_target, due=due, g=g)
    except Exception:
        result = None

    st.divider()
    if err:
        st.error(err)
    elif result is None:
        st.error("Couldn't compute a result with these inputs — try adjusting the values (e.g. growth rate close to the discount rate can be sensitive).")
    else:
        timing_word = "at the start" if due else "at the end"
        growth_phrase = f", growing {g_pct:.2f}% each period" if is_growing else ""
        if solve_for == "Future Value (FV)":
            st.metric("Future Value (FV)", fmt_money_plain(result))
            st.success(f"Paying {fmt_money(pmt)} {timing_word} of each period{growth_phrase} for {n:.0f} periods "
                       f"at {rate_pct:.2f}% per period grows to **{fmt_money(result)}**.")
            n_int = max(1, int(round(n)))
            flows = {t: pmt * (1 + g) ** (t - 1) for t in range(1, n_int + 1)}
        elif solve_for == "Present Value (PV)":
            st.metric("Present Value (PV)", fmt_money_plain(result))
            st.success(f"A stream of {fmt_money(pmt)} paid {timing_word} of each period{growth_phrase} for "
                       f"{n:.0f} periods is worth **{fmt_money(result)}** today at a discount rate of {rate_pct:.2f}%.")
            n_int = max(1, int(round(n)))
            flows = {t: pmt * (1 + g) ** (t - 1) for t in range(1, n_int + 1)}
            flows[0] = -result
        elif solve_for == "Payment (PMT)":
            st.metric("Payment (PMT)" + (" — first payment" if is_growing else ""), fmt_money_plain(result))
            target_desc = f"a present value of {fmt_money(pv_target)}" if reference == "Present Value (PV)" else f"a future value of {fmt_money(fv_target)}"
            st.success(f"To reach {target_desc} over {n:.0f} periods at {rate_pct:.2f}% per period, "
                       f"you'd need to pay **{fmt_money(result)}** {timing_word} of each period{growth_phrase}.")
            n_int = max(1, int(round(n)))
            flows = {t: result * (1 + g) ** (t - 1) for t in range(1, n_int + 1)}
            if reference == "Present Value (PV)":
                flows[0] = -pv_target
        elif solve_for == "Number of Periods (N)":
            st.metric("Number of Periods (N)", f"{result:.2f}")
            target_desc = f"{fmt_money(pv_target)} today" if reference == "Present Value (PV)" else f"{fmt_money(fv_target)} in the future"
            st.success(f"Paying {fmt_money(pmt)} {timing_word} of each period{growth_phrase} at {rate_pct:.2f}% "
                       f"per period, it takes about **{result:.2f} periods** to reach {target_desc}.")
            n_int = max(1, int(round(result)))
            flows = {t: pmt * (1 + g) ** (t - 1) for t in range(1, n_int + 1)}
        else:
            st.metric("Rate per Period (I/Y)", fmt_pct(result))
            target_desc = f"{fmt_money(pv_target)} today" if reference == "Present Value (PV)" else f"{fmt_money(fv_target)} in the future"
            st.success(f"Paying {fmt_money(pmt)} {timing_word} of each period{growth_phrase} for {n:.0f} periods "
                       f"to reach {target_desc} implies a periodic rate of **{fmt_pct(result)}**.")
            n_int = max(1, int(round(n)))
            flows = {t: pmt * (1 + g) ** (t - 1) for t in range(1, n_int + 1)}

        with st.expander("📐 Formula used"):
            if not is_growing:
                st.code("PV = PMT × [1 − (1+i)⁻ⁿ] ÷ i        FV = PMT × [(1+i)ⁿ − 1] ÷ i\n(× (1+i) for annuity due)")
            else:
                st.code("PV = PMT₁ ÷ (i−g) × [1 − ((1+g)/(1+i))ⁿ]     FV = PMT₁ × [(1+i)ⁿ − (1+g)ⁿ] ÷ (i−g)\n(× (1+i) for annuity due)")

        st.caption("Cash-flow timeline:")
        chart_df = pd.DataFrame({"Cash Flow": flows}).sort_index()
        st.bar_chart(chart_df)

# ============================================================
# TAB: PERPETUITY
# ============================================================
with tab_perp:
    st.subheader("♾️ Perpetuity")
    st.write("A stream of equal (or steadily growing) payments that continues forever — e.g. preferred stock dividends or an endowment fund.")

    is_growing_p = st.checkbox("Payments grow each period (Gordon growth model)", key="pp_growing")
    solve_for_p = st.selectbox("Solve for:", ["Present Value (PV)", "Payment (PMT)", "Rate per Period (I/Y)"], key="pp_solve")

    c1, c2 = st.columns(2)
    pmt_p = rate_pct_p = pv_p = g_pct_p = None
    with c1:
        if solve_for_p != "Payment (PMT)":
            pmt_p = st.number_input("Payment per period (PMT)" + (" — first payment" if is_growing_p else ""),
                                     value=50.0, step=5.0, key="pp_pmt")
        if is_growing_p:
            g_pct_p = st.number_input("Growth Rate per Period (%)", value=2.0, step=0.25, key="pp_g")
        else:
            g_pct_p = 0.0
    with c2:
        if solve_for_p != "Rate per Period (I/Y)":
            rate_pct_p = st.number_input("Rate per Period, I/Y (%)", value=8.0, step=0.25, key="pp_rate")
        if solve_for_p != "Present Value (PV)":
            pv_p = st.number_input("Present Value (PV)", value=625.0, step=25.0, key="pp_pv")

    result_p, err_p = None, None
    g_p = (g_pct_p or 0.0) / 100
    if solve_for_p == "Present Value (PV)":
        if rate_pct_p / 100 <= g_p:
            err_p = "The discount rate must be greater than the growth rate, or the value is infinite."
        else:
            result_p = pv_perpetuity(pmt_p, rate_pct_p / 100, g=g_p)
    elif solve_for_p == "Payment (PMT)":
        if rate_pct_p / 100 <= g_p:
            err_p = "The discount rate must be greater than the growth rate."
        else:
            result_p = pmt_from_perpetuity(pv_p, rate_pct_p / 100, g=g_p)
    else:
        result_p = rate_from_perpetuity(pv_p, pmt_p, g=g_p)

    st.divider()
    if err_p:
        st.error(err_p)
    elif result_p is None:
        st.error("Couldn't compute a result — check your inputs.")
    else:
        growth_phrase_p = f", growing {g_pct_p:.2f}% per period" if is_growing_p else ""
        if solve_for_p == "Present Value (PV)":
            st.metric("Present Value (PV)", fmt_money_plain(result_p))
            st.success(f"A perpetual stream starting at {fmt_money(pmt_p)} per period{growth_phrase_p}, "
                       f"discounted at {rate_pct_p:.2f}%, is worth **{fmt_money(result_p)}** today.")
            formula_p = "PV = PMT ÷ (i − g)" if is_growing_p else "PV = PMT ÷ i"
        elif solve_for_p == "Payment (PMT)":
            st.metric("Payment (PMT)", fmt_money_plain(result_p))
            st.success(f"A perpetuity worth {fmt_money(pv_p)} today at a discount rate of {rate_pct_p:.2f}%{growth_phrase_p} "
                       f"would need to pay **{fmt_money(result_p)}** per period (first payment).")
            formula_p = "PMT = PV × (i − g)" if is_growing_p else "PMT = PV × i"
        else:
            st.metric("Rate per Period (I/Y)", fmt_pct(result_p))
            st.success(f"A perpetuity worth {fmt_money(pv_p)} paying {fmt_money(pmt_p)} per period{growth_phrase_p} "
                       f"implies a discount rate of **{fmt_pct(result_p)}**.")
            formula_p = "i = PMT ÷ PV + g"

        with st.expander("📐 Formula used"):
            st.code(formula_p)
        st.caption("💡 Real-world context: valuing preferred stock (level dividend) or using the Gordon Growth Model to value common stock (growing dividend), or sizing an endowment/scholarship fund.")

# ============================================================
# TAB: APR <-> EAR
# ============================================================
FREQ_OPTIONS = {"Annually": 1, "Semiannually": 2, "Quarterly": 4, "Monthly": 12, "Daily": 365, "Continuous": None}

with tab_aprear:
    st.subheader("🔄 APR ↔ EAR Converter")
    st.write("The stated annual rate (APR) understates your true annual cost or return whenever compounding "
             "happens more than once a year. The Effective Annual Rate (EAR / APY) shows the real number.")

    direction = st.radio("I know the:", ["APR (stated/nominal rate)", "EAR (effective annual rate)"],
                          key="ae_dir", horizontal=True)
    c1, c2 = st.columns(2)
    with c1:
        rate_in = st.number_input(f"{'APR' if direction.startswith('APR') else 'EAR'} (%)", value=6.0, step=0.25, key="ae_rate")
    with c2:
        freq_label = st.selectbox("Compounding frequency", list(FREQ_OPTIONS.keys()), index=3, key="ae_freq")
    m = FREQ_OPTIONS[freq_label]

    if direction.startswith("APR"):
        ear = ear_from_apr(rate_in / 100, m)
        st.metric("Effective Annual Rate (EAR)", fmt_pct(ear))
        st.success(f"A stated rate of {rate_in:.2f}% compounded {freq_label.lower()} is actually "
                   f"**{fmt_pct(ear)}** per year once compounding is accounted for.")
    else:
        apr = apr_from_ear(rate_in / 100, m)
        st.metric("APR (nominal rate)", fmt_pct(apr))
        st.success(f"To achieve an effective annual rate of {rate_in:.2f}% with {freq_label.lower()} compounding, "
                   f"the stated (nominal) rate would need to be **{fmt_pct(apr)}**.")

    with st.expander("📊 Compare EAR across compounding frequencies"):
        base_apr = rate_in / 100 if direction.startswith("APR") else apr_from_ear(rate_in / 100, m)
        comp_rows = []
        for label, freq in FREQ_OPTIONS.items():
            comp_rows.append({"Compounding": label, "EAR": ear_from_apr(base_apr, freq)})
        comp_df = pd.DataFrame(comp_rows)
        st.dataframe(comp_df.assign(EAR=comp_df["EAR"].apply(fmt_pct)), use_container_width=True, hide_index=True)
        st.bar_chart(comp_df.set_index("Compounding")["EAR"])
        st.caption(f"All rows show the EAR for the same {fmt_pct(base_apr)} stated annual rate — "
                   f"only the compounding frequency changes. More frequent compounding → higher EAR.")

    with st.expander("🧮 Bonus: Real vs. Nominal Rate (Fisher Equation)"):
        st.write("Inflation erodes purchasing power — the Fisher equation finds your *real* return after inflation.")
        fc1, fc2 = st.columns(2)
        with fc1:
            nominal_in = st.number_input("Nominal rate (%)", value=7.0, step=0.25, key="fisher_nom")
        with fc2:
            inflation_in = st.number_input("Inflation rate (%)", value=3.0, step=0.25, key="fisher_inf")
        real_exact = fisher_real_rate(nominal_in / 100, inflation_in / 100)
        real_approx = nominal_in / 100 - inflation_in / 100
        rc1, rc2 = st.columns(2)
        rc1.metric("Real rate (exact Fisher equation)", fmt_pct(real_exact))
        rc2.metric("Real rate (quick approximation)", fmt_pct(real_approx))
        st.caption("Exact: (1 + real) = (1 + nominal) ÷ (1 + inflation). The simple subtraction "
                   "(nominal − inflation) is a common shortcut that's close at low rates but drifts "
                   "apart as rates rise.")

# ============================================================
# TAB: MORTGAGE / LOAN
# ============================================================
with tab_mortgage:
    st.subheader("🏠 Mortgage / Loan Amortization")
    mortgage_mode = st.radio("I'm...", ["Buying a home (new mortgage)", "Refinancing an existing mortgage"],
                              key="mtg_mode", horizontal=True)
    is_refi = mortgage_mode.startswith("Refinancing")

    include_pmi, pmi_rate = False, 0.0
    refi_summary_rows, refi_note = [], None

    if not is_refi:
        st.write("Enter the home price and down payment to see your loan amount, payment, a full amortization schedule, and how extra payments accelerate payoff.")

        st.markdown("**Home Price & Down Payment**")
        hp1, hp2, hp3 = st.columns(3)
        with hp1:
            home_price = st.number_input("Home Price", value=437500.0, step=5000.0, min_value=0.0, key="mtg_price")
            down_mode = st.radio("Down payment as", ["Percent of price", "Dollar amount"], key="mtg_down_mode", horizontal=True)
        with hp2:
            if down_mode == "Percent of price":
                down_pct_in = st.number_input("Down Payment (%)", value=20.0, step=1.0, min_value=0.0, max_value=100.0, key="mtg_down_pct")
                down_payment = home_price * down_pct_in / 100
            else:
                down_payment = st.number_input("Down Payment ($)", value=87500.0, step=1000.0, min_value=0.0, key="mtg_down_amt")
                down_pct_in = (down_payment / home_price * 100) if home_price > 0 else 0.0
        with hp3:
            loan_amount = max(home_price - down_payment, 0.0)
            ltv = (loan_amount / home_price * 100) if home_price > 0 else 0.0
            st.metric("Loan Amount (financed)", fmt_money_plain(loan_amount))
            st.caption(f"Down payment: {fmt_money_plain(down_payment)} ({down_pct_in:.1f}%) · Loan-to-Value: {ltv:.1f}%")

        if 0 < loan_amount and ltv > 80:
            st.warning(f"⚠️ With a down payment under 20% (LTV {ltv:.1f}%), lenders typically require **Private "
                       f"Mortgage Insurance (PMI)** until you reach 20% equity. You can estimate that cost below.")
            pmi_c1, pmi_c2 = st.columns(2)
            with pmi_c1:
                include_pmi = st.checkbox("Include an estimated PMI cost", key="mtg_pmi_toggle")
            with pmi_c2:
                pmi_rate = st.number_input("Estimated PMI Rate (annual, % of loan)", value=0.75, step=0.05,
                                            key="mtg_pmi_rate", disabled=not include_pmi)

        st.markdown("**Loan Terms**")
        mc1, mc2, mc3 = st.columns(3)
        with mc1:
            annual_rate_m = st.number_input("Annual Interest Rate (%)", value=6.5, step=0.125, key="mtg_rate")
        with mc2:
            years_m = st.number_input("Loan Term (years)", value=30.0, step=1.0, min_value=1.0, key="mtg_years")
            freq_label_m = st.selectbox("Payments per Year", ["Monthly (12)", "Biweekly (26)", "Weekly (52)", "Annual (1)"],
                                         index=0, key="mtg_freq")
        with mc3:
            ppy_m = {"Monthly (12)": 12, "Biweekly (26)": 26, "Weekly (52)": 52, "Annual (1)": 1}[freq_label_m]
            extra_pmt = st.number_input("Extra payment per period (optional)", value=0.0, step=50.0, key="mtg_extra")

    else:
        st.write("Compare staying with your current loan vs. refinancing into a new one — including a break-even "
                 "estimate for the closing costs.")

        st.markdown("**Current Loan (what you have now)**")
        cur1, cur2, cur3 = st.columns(3)
        with cur1:
            current_balance = st.number_input("Current Remaining Balance", value=320000.0, step=5000.0, min_value=0.0, key="refi_balance")
        with cur2:
            current_rate = st.number_input("Current Interest Rate (%)", value=7.25, step=0.125, key="refi_current_rate")
        with cur3:
            years_remaining = st.number_input("Years Remaining on Current Loan", value=27.0, step=1.0, min_value=0.5, key="refi_years_remaining")

        st.markdown("**New Loan (what you're considering)**")
        new1, new2, new3 = st.columns(3)
        with new1:
            new_rate = st.number_input("New Interest Rate (%)", value=6.25, step=0.125, key="refi_new_rate")
        with new2:
            new_term = st.number_input("New Loan Term (years)", value=30.0, step=1.0, min_value=1.0, key="refi_new_term")
        with new3:
            closing_costs = st.number_input("Estimated Closing Costs / Fees", value=6000.0, step=500.0, min_value=0.0, key="refi_costs")

        rc1, rc2 = st.columns(2)
        with rc1:
            roll_in = st.checkbox("Roll closing costs into the new loan (instead of paying upfront)", key="refi_roll_in")
        with rc2:
            stay_years = st.number_input("Years you plan to stay in this home (optional, for a recommendation)",
                                          value=10.0, step=1.0, min_value=0.0, key="refi_stay_years")

        st.caption("Note: this tool assumes a standard rate-and-term refinance (monthly payments, no extra payments, "
                   "no PMI change) to keep the comparison focused — switch to \"Buying\" mode for those options.")

        current_schedule, current_summary = amortization_schedule(current_balance, current_rate / 100, years_remaining, 12, 0.0)
        new_principal = current_balance + (closing_costs if roll_in else 0.0)

        current_monthly = current_summary["scheduled_payment"]
        current_total_cost = current_monthly * current_summary["n_actual"]
        monthly_savings = None  # filled in after new_summary is computed below

        st.divider()
        st.markdown("#### 📊 Refinance Analysis")

        # These become the shared "going forward" loan for the rest of the tab (schedule, charts, downloads)
        loan_amount, annual_rate_m, years_m, ppy_m, extra_pmt = new_principal, new_rate, new_term, 12, 0.0

    if loan_amount <= 0:
        st.error("Loan amount is $0 — there's nothing to amortize with these inputs.")
    else:
        schedule_m, summary_m = amortization_schedule(loan_amount, annual_rate_m / 100, years_m, ppy_m, extra_pmt)
        schedule_no_extra, summary_no_extra = amortization_schedule(loan_amount, annual_rate_m / 100, years_m, ppy_m, 0.0)
        monthly_pmi = (loan_amount * pmi_rate / 100 / ppy_m) if include_pmi else 0.0

        if is_refi:
            new_monthly = summary_m["scheduled_payment"]
            new_total_cost = new_monthly * summary_m["n_actual"] + (0.0 if roll_in else closing_costs)
            monthly_savings = current_monthly - new_monthly
            lifetime_savings = current_total_cost - new_total_cost

            rm1, rm2, rm3 = st.columns(3)
            rm1.metric("Current Payment", fmt_money_plain(current_monthly))
            rm2.metric("New Payment", fmt_money_plain(new_monthly), delta=fmt_money_plain(-monthly_savings) if monthly_savings else None,
                       delta_color="inverse")
            rm3.metric("Monthly Savings" if monthly_savings >= 0 else "Monthly Increase", fmt_money_plain(abs(monthly_savings)))

            breakeven_months = (closing_costs / monthly_savings) if (monthly_savings and monthly_savings > 0 and not roll_in) else None

            if roll_in:
                st.info(f"Closing costs of {fmt_money(closing_costs)} are rolled into the new loan balance, so there's no "
                        f"separate amount to recoup — compare the **lifetime cost** below instead of a break-even point.")
            elif breakeven_months is not None:
                years_txt = f"{breakeven_months / 12:.1f} years"
                if stay_years > 0 and stay_years * 12 >= breakeven_months:
                    st.success(f"Refinancing saves {fmt_money(monthly_savings)}/month. Paying {fmt_money(closing_costs)} in "
                               f"closing costs upfront, you'd **break even in about {breakeven_months:.0f} months ({years_txt})** "
                               f"— since you plan to stay {stay_years:.0f} years, that's **worth it**.")
                elif stay_years > 0:
                    st.warning(f"Refinancing saves {fmt_money(monthly_savings)}/month, but breaking even on the "
                               f"{fmt_money(closing_costs)} closing costs takes about **{breakeven_months:.0f} months ({years_txt})** "
                               f"— longer than the {stay_years:.0f} years you plan to stay, so it **may not pay off** before you'd move.")
                else:
                    st.info(f"Refinancing saves {fmt_money(monthly_savings)}/month. You'd break even on the "
                            f"{fmt_money(closing_costs)} closing costs in about **{breakeven_months:.0f} months ({years_txt})**.")
            else:
                st.info("This new rate/term doesn't lower your monthly payment — but a shorter term can still reduce "
                        "total interest paid over time. Compare the lifetime cost below.")

            lc1, lc2 = st.columns(2)
            lc1.metric(f"Total Cost, Staying (remaining {years_remaining:.0f} yrs)", fmt_money_plain(current_total_cost))
            lc2.metric(f"Total Cost, Refinancing ({new_term:.0f} yrs)", fmt_money_plain(new_total_cost))
            if lifetime_savings > 0:
                st.caption(f"Refinancing costs **{fmt_money(abs(lifetime_savings))} less** over the compared horizons above "
                           f"(note the terms differ — {years_remaining:.0f} vs {new_term:.0f} years — so this isn't a purely "
                           f"apples-to-apples comparison, just what you'd actually pay under each path).")
            else:
                st.caption(f"Refinancing costs **{fmt_money(abs(lifetime_savings))} more** over the compared horizons above "
                           f"(note the terms differ — {years_remaining:.0f} vs {new_term:.0f} years).")

            refi_summary_rows = [
                ("Current Balance", fmt_money_plain(current_balance)), ("Current Rate", fmt_pct(current_rate / 100)),
                ("Years Remaining (current)", f"{years_remaining:.1f}"), ("Current Payment", fmt_money_plain(current_monthly)),
                ("New Rate", fmt_pct(new_rate / 100)), ("New Term (years)", f"{new_term:.0f}"),
                ("Closing Costs", fmt_money_plain(closing_costs)), ("Rolled Into Loan", "Yes" if roll_in else "No"),
                ("New Payment", fmt_money_plain(new_monthly)),
                ("Monthly Savings", fmt_money_plain(monthly_savings)),
                ("Break-even (months)", f"{breakeven_months:.0f}" if breakeven_months is not None else "N/A"),
            ]

            st.divider()
            st.markdown("#### 📋 New Loan Details")

        r1, r2, r3 = st.columns(3)
        r1.metric("Scheduled Payment (Principal & Interest)", fmt_money_plain(summary_m["scheduled_payment"]))
        r2.metric("Total Interest Paid", fmt_money_plain(summary_m["total_interest"]))
        r3.metric("Payoff Time", f"{summary_m['payoff_years']:.1f} years")

        if include_pmi and monthly_pmi > 0:
            st.info(f"Estimated PMI adds about {fmt_money(monthly_pmi)} per period — "
                    f"**{fmt_money(summary_m['scheduled_payment'] + monthly_pmi)}** total per period including PMI. "
                    f"PMI typically drops off automatically once you reach 20% equity.")

        if extra_pmt > 0:
            years_saved = summary_no_extra["payoff_years"] - summary_m["payoff_years"]
            interest_saved = summary_no_extra["total_interest"] - summary_m["total_interest"]
            st.success(f"Paying an extra {fmt_money(extra_pmt)} per period pays off the loan "
                       f"**{years_saved:.1f} years early** and saves **{fmt_money(interest_saved)}** in interest.")
        else:
            st.info(f"Over the life of this loan you'll pay {fmt_money(summary_m['total_interest'])} in interest — "
                    f"that's {summary_m['total_interest'] / loan_amount * 100:.1f}% of the amount borrowed.")

        df_sched = pd.DataFrame(schedule_m)
        st.markdown("**Balance over time**")
        st.line_chart(df_sched.set_index("period")[["balance"]])

        st.markdown("**Cumulative interest vs. principal paid**")
        df_sched["cum_interest"] = df_sched["interest"].cumsum()
        df_sched["cum_principal"] = df_sched["principal"].cumsum()
        st.area_chart(df_sched.set_index("period")[["cum_principal", "cum_interest"]])

        with st.expander("📋 Full amortization schedule"):
            st.dataframe(
                df_sched[["period", "payment", "interest", "principal", "balance"]].style.format(
                    {"payment": "${:,.2f}", "interest": "${:,.2f}", "principal": "${:,.2f}", "balance": "${:,.2f}"}
                ),
                use_container_width=True, height=350,
            )

        st.markdown("#### ⬇️ Download")
        d1, d2 = st.columns(2)
        with d1:
            csv_bytes = df_sched.to_csv(index=False).encode("utf-8")
            st.download_button("CSV", data=csv_bytes, file_name="amortization_schedule.csv", mime="text/csv", use_container_width=True)
        with d2:
            if is_refi:
                summary_rows_m = refi_summary_rows + [
                    ("Payments per Year", ppy_m), ("Total Interest (new loan)", fmt_money_plain(summary_m["total_interest"])),
                    ("Payoff Time (years)", f"{summary_m['payoff_years']:.2f}"),
                ]
                excel_title = "Mortgage Refinance Analysis"
            else:
                summary_rows_m = [
                    ("Home Price", fmt_money_plain(home_price)), ("Down Payment", f"{fmt_money_plain(down_payment)} ({down_pct_in:.1f}%)"),
                    ("Loan Amount", fmt_money_plain(loan_amount)), ("Loan-to-Value (LTV)", f"{ltv:.1f}%"),
                    ("Annual Rate", fmt_pct(annual_rate_m / 100)), ("Term (years)", f"{years_m:.0f}"),
                    ("Payments per Year", ppy_m), ("Extra Payment", fmt_money_plain(extra_pmt)),
                    ("Scheduled Payment (P&I)", fmt_money_plain(summary_m["scheduled_payment"])),
                    ("Total Interest", fmt_money_plain(summary_m["total_interest"])), ("Payoff Time (years)", f"{summary_m['payoff_years']:.2f}"),
                ]
                if include_pmi:
                    summary_rows_m += [("Estimated PMI Rate", fmt_pct(pmi_rate / 100)), ("Estimated PMI per Period", fmt_money_plain(monthly_pmi))]
                excel_title = "Mortgage / Loan Amortization Schedule"
            excel_buf_m = build_amortization_excel(excel_title, summary_rows_m, schedule_m)
            st.download_button("Excel (.xlsx)", data=excel_buf_m, file_name="amortization_schedule.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

        pdf_table_rows = [[r["period"], fmt_money_plain(r["payment"]), fmt_money_plain(r["interest"]), fmt_money_plain(r["principal"]), fmt_money_plain(r["balance"])]
                           for r in schedule_m]
        if is_refi:
            pdf_subtitle = (f"Current: {fmt_money_plain(current_balance)} at {current_rate:.2f}%  |  "
                             f"New: {fmt_money_plain(loan_amount)} at {annual_rate_m:.2f}% for {years_m:.0f} years")
            pdf_title = "Mortgage Refinance Analysis"
        else:
            pdf_subtitle = (f"Home Price: {fmt_money_plain(home_price)}  |  Down: {fmt_money_plain(down_payment)} ({down_pct_in:.1f}%)  |  "
                             f"Loan: {fmt_money_plain(loan_amount)} at {annual_rate_m:.2f}% for {years_m:.0f} years")
            pdf_title = "Mortgage / Loan Amortization Schedule"
        pdf_buf_m = build_simple_pdf(
            pdf_title, pdf_subtitle, summary_rows_m, ["Period", "Payment", "Interest", "Principal", "Balance"], pdf_table_rows,
            footnote="Generated by the Time Value of Money Calculator.",
        )
        st.download_button("⬇️ PDF Report", data=pdf_buf_m, file_name="amortization_schedule.pdf", mime="application/pdf")

# ============================================================
# TAB: RETIREMENT PLANNER
# ============================================================
with tab_retire:
    st.subheader("🎯 Retirement Planner")
    st.write("Projects your savings to retirement, accounting for investment growth, contributions that grow with "
             "your salary, and inflation eroding purchasing power along the way.")

    rc1, rc2, rc3 = st.columns(3)
    with rc1:
        current_age = st.number_input("Current Age", value=30, step=1, key="ret_age")
        retire_age = st.number_input("Retirement Age", value=65, step=1, key="ret_retage")
        current_savings = st.number_input("Current Savings", value=25000.0, step=1000.0, key="ret_savings")
    with rc2:
        annual_contribution = st.number_input("Current Annual Contribution", value=6000.0, step=500.0, key="ret_contrib")
        contrib_growth = st.number_input("Expected Contribution Growth (salary growth), %/yr", value=3.0, step=0.25, key="ret_cgrowth")
        return_rate = st.number_input("Expected Annual Investment Return (%)", value=7.0, step=0.25, key="ret_return")
    with rc3:
        inflation_rate = st.number_input("Expected Inflation Rate (%)", value=3.0, step=0.25, key="ret_inflation")
        goal_type = st.radio("Goal type", ["Target nest egg (today's $)", "Target annual income (today's $)"], key="ret_goaltype")
        if goal_type.startswith("Target nest"):
            target_today = st.number_input("Target Nest Egg (in today's dollars)", value=1500000.0, step=50000.0, key="ret_target_egg")
            withdrawal_rate = None
        else:
            target_today = st.number_input("Target Annual Retirement Income (today's dollars)", value=60000.0, step=2000.0, key="ret_target_income")
            withdrawal_rate = st.number_input("Assumed Safe Withdrawal Rate (%)", value=4.0, step=0.25, key="ret_wd")

    n_years = max(int(retire_age) - int(current_age), 0)
    if n_years <= 0:
        st.error("Retirement age must be after current age.")
    else:
        i_r, g_r, infl_r = return_rate / 100, contrib_growth / 100, inflation_rate / 100
        rows_r = growing_annuity_projection(current_savings, annual_contribution, i_r, g_r, n_years)
        nominal_fv = rows_r[-1]["balance"]
        real_fv = nominal_fv / (1 + infl_r) ** n_years

        if goal_type.startswith("Target nest"):
            target_nominal = target_today * (1 + infl_r) ** n_years
        else:
            required_egg_today = target_today / (withdrawal_rate / 100)
            target_nominal = required_egg_today * (1 + infl_r) ** n_years
        target_real = target_nominal / (1 + infl_r) ** n_years  # == target_today, shown for symmetry

        gap_nominal = nominal_fv - target_nominal

        st.divider()
        m1, m2, m3 = st.columns(3)
        m1.metric("Projected Balance at Retirement (nominal $)", fmt_money_plain(nominal_fv))
        m2.metric("Same Balance in Today's Dollars (real)", fmt_money_plain(real_fv))
        m3.metric("Target Needed (nominal $ at retirement)", fmt_money_plain(target_nominal))

        if gap_nominal >= 0:
            st.success(f"🎉 On track! Your projected balance exceeds your target by **{fmt_money(gap_nominal)}** "
                       f"(in future dollars), equivalent to {fmt_money(gap_nominal / (1+infl_r)**n_years)} in today's purchasing power.")
        else:
            fv_existing_alone = fv_single(current_savings, i_r, n_years)
            needed_from_contrib = target_nominal - fv_existing_alone
            if needed_from_contrib <= 0:
                st.warning("Your current savings alone should reach the target — but double check your assumptions.")
            else:
                required_pmt1 = pmt1_from_growing_fv(needed_from_contrib, i_r, n_years, g_r)
                st.warning(f"⚠️ Shortfall of **{fmt_money(-gap_nominal)}** (future dollars) vs. your target. "
                           f"To close the gap, your first-year contribution would need to be about "
                           f"**{fmt_money(required_pmt1)}** (vs. {fmt_money(annual_contribution)} planned), "
                           f"still growing {contrib_growth:.1f}%/year.")

        df_r = pd.DataFrame(rows_r)
        df_r["age"] = df_r["year"] + current_age
        df_r["real_balance"] = df_r["balance"] / (1 + infl_r) ** df_r["year"]

        st.markdown("**Balance over time: nominal vs. today's-dollar (real) value**")
        st.line_chart(df_r.set_index("age")[["balance", "real_balance"]].rename(
            columns={"balance": "Nominal Balance", "real_balance": "Real (Today's $) Balance"}))

        with st.expander("📋 Year-by-year projection"):
            st.dataframe(
                df_r[["age", "contribution", "balance", "real_balance"]].rename(
                    columns={"age": "Age", "contribution": "Contribution", "balance": "Nominal Balance", "real_balance": "Real Balance"}
                ).style.format({"Contribution": "${:,.2f}", "Nominal Balance": "${:,.2f}", "Real Balance": "${:,.2f}"}),
                use_container_width=True, height=350,
            )

        st.markdown("#### ⬇️ Download")
        d1, d2 = st.columns(2)
        with d1:
            csv_r = df_r.to_csv(index=False).encode("utf-8")
            st.download_button("CSV", data=csv_r, file_name="retirement_projection.csv", mime="text/csv", use_container_width=True)
        with d2:
            summary_rows_r = [
                ("Current Age", current_age), ("Retirement Age", retire_age), ("Years to Grow", n_years),
                ("Current Savings", fmt_money_plain(current_savings)), ("Annual Contribution (Year 1)", fmt_money_plain(annual_contribution)),
                ("Contribution Growth Rate", fmt_pct(g_r)), ("Investment Return", fmt_pct(i_r)),
                ("Inflation Rate", fmt_pct(infl_r)), ("Projected Nominal Balance", fmt_money_plain(nominal_fv)),
                ("Projected Real (Today's $) Balance", fmt_money_plain(real_fv)), ("Target (nominal $ at retirement)", fmt_money_plain(target_nominal)),
            ]
            schedule_for_excel = [dict(period=r["age"], payment=r["contribution"], interest=0, principal=r["contribution"], balance=r["balance"]) for r in df_r.to_dict("records")]
            excel_buf_r = build_amortization_excel("Retirement Savings Projection", summary_rows_r, schedule_for_excel)
            st.download_button("Excel (.xlsx)", data=excel_buf_r, file_name="retirement_projection.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

        pdf_rows_r = [[int(r["age"]), fmt_money_plain(r["contribution"]), fmt_money_plain(r["balance"]), fmt_money_plain(r["real_balance"])] for r in df_r.to_dict("records")]
        pdf_buf_r = build_simple_pdf(
            "Retirement Savings Projection", f"Age {current_age} → {retire_age}  |  Return {return_rate:.1f}%, Inflation {inflation_rate:.1f}%",
            summary_rows_r, ["Age", "Contribution", "Nominal Balance", "Real Balance"], pdf_rows_r,
            footnote="Nominal = actual future dollars. Real = equivalent purchasing power in today's dollars. "
                     "Generated by the Time Value of Money Calculator.",
        )
        st.download_button("⬇️ PDF Report", data=pdf_buf_r, file_name="retirement_projection.pdf", mime="application/pdf")

# ============================================================
# TAB: LOTTERY
# ============================================================
with tab_lottery:
    st.subheader("🎰 Lottery: Lump Sum vs. Annuity")
    st.write("Big jackpots are typically paid as ~30 annual payments that grow ~5%/year, or you can take a "
             "smaller lump sum today. Which is worth more depends on your discount rate assumption.")

    lc1, lc2, lc3 = st.columns(3)
    with lc1:
        first_payment = st.number_input("First Annuity Payment", value=10_000_000.0, step=500_000.0, key="lot_first")
        n_payments = st.number_input("Number of Payments", value=30, step=1, key="lot_n")
    with lc2:
        payment_growth = st.number_input("Annual Payment Growth (%)", value=5.0, step=0.25, key="lot_growth")
        discount_rate_l = st.number_input("Your Discount Rate / Expected Return (%)", value=6.0, step=0.25, key="lot_disc")
    with lc3:
        lump_sum = st.number_input("Lump Sum Cash Value Offered", value=180_000_000.0, step=1_000_000.0, key="lot_lump")
        apply_tax = st.checkbox("Apply a flat combined tax rate", key="lot_tax_toggle")
        tax_rate = st.number_input("Combined Tax Rate (%)", value=37.0, step=1.0, key="lot_tax", disabled=not apply_tax)

    pv_annuity_opt = lottery_pv_of_annuity(first_payment, int(n_payments), payment_growth / 100, discount_rate_l / 100)
    breakeven = lottery_breakeven_rate(first_payment, int(n_payments), payment_growth / 100, lump_sum)

    tax_mult = (1 - tax_rate / 100) if apply_tax else 1.0
    pv_annuity_after = pv_annuity_opt * tax_mult
    lump_after = lump_sum * tax_mult

    st.divider()
    m1, m2 = st.columns(2)
    m1.metric(f"PV of Annuity Option{' (after tax)' if apply_tax else ''}", fmt_money_plain(pv_annuity_after))
    m2.metric(f"Lump Sum{' (after tax)' if apply_tax else ''}", fmt_money_plain(lump_after))

    if pv_annuity_after > lump_after:
        st.success(f"At a {discount_rate_l:.2f}% discount rate, the **annuity option is worth "
                   f"{fmt_money(pv_annuity_after - lump_after)} more** in present-value terms.")
    else:
        st.success(f"At a {discount_rate_l:.2f}% discount rate, the **lump sum is worth "
                   f"{fmt_money(lump_after - pv_annuity_after)} more** in present-value terms.")

    if breakeven is not None:
        st.info(f"**Break-even discount rate: {fmt_pct(breakeven)}.** Below this rate, the annuity's present "
                f"value is higher; above it, the lump sum wins. (Taxes applied equally to both don't change this "
                f"break-even point.)")

    rate_range = [r / 1000 for r in range(0, 121, 2)]  # 0% to 12% in 0.2% steps
    sens_df = pd.DataFrame({
        "Discount Rate": [f"{r*100:.1f}%" for r in rate_range],
        "PV of Annuity": [lottery_pv_of_annuity(first_payment, int(n_payments), payment_growth / 100, r) for r in rate_range],
        "Lump Sum": [lump_sum] * len(rate_range),
    })
    st.markdown("**Sensitivity: PV of annuity option vs. discount rate (lump sum shown as flat reference line)**")
    st.line_chart(sens_df.set_index("Discount Rate")[["PV of Annuity", "Lump Sum"]])

    st.caption("Simplifications: taxes are modeled as a flat combined rate applied equally to both options "
               "(real tax treatment is more complex); annuity payments are assumed to start immediately (annuity due).")

    summary_rows_l = [
        ("First Payment", fmt_money_plain(first_payment)), ("Number of Payments", int(n_payments)),
        ("Payment Growth Rate", fmt_pct(payment_growth / 100)), ("Discount Rate", fmt_pct(discount_rate_l / 100)),
        ("PV of Annuity Option", fmt_money_plain(pv_annuity_opt)), ("Lump Sum Offered", fmt_money_plain(lump_sum)),
        ("Break-even Discount Rate", fmt_pct(breakeven) if breakeven else "N/A"),
    ]
    if apply_tax:
        summary_rows_l += [("Tax Rate Applied", fmt_pct(tax_rate / 100)),
                            ("PV of Annuity (after tax)", fmt_money_plain(pv_annuity_after)),
                            ("Lump Sum (after tax)", fmt_money_plain(lump_after))]

    st.markdown("#### ⬇️ Download")
    d1, d2 = st.columns(2)
    with d1:
        csv_l = sens_df.to_csv(index=False).encode("utf-8")
        st.download_button("CSV (sensitivity table)", data=csv_l, file_name="lottery_comparison.csv", mime="text/csv", use_container_width=True)
    with d2:
        pdf_buf_l = build_simple_pdf("Lottery: Lump Sum vs. Annuity", "Comparison summary", summary_rows_l,
                                      footnote="Generated by the Time Value of Money Calculator.")
        st.download_button("PDF Report", data=pdf_buf_l, file_name="lottery_comparison.pdf", mime="application/pdf", use_container_width=True)

# ============================================================
# TAB: CREDIT CARD
# ============================================================
with tab_cc:
    st.subheader("💳 Credit Card: APR vs. EAR")
    st.write("Credit cards advertise an APR but typically compound daily — here's what that really costs you.")

    cc1, cc2 = st.columns(2)
    with cc1:
        cc_apr = st.number_input("Card's Stated APR (%)", value=24.99, step=0.5, key="cc_apr")
        cc_balance = st.number_input("Balance Carried", value=3000.0, step=100.0, key="cc_balance")
    with cc2:
        cc_freq_label = st.selectbox("Compounding Frequency", list(FREQ_OPTIONS.keys()), index=4, key="cc_freq")
        cc_months = st.number_input("Months with No Payments (illustration)", value=12, step=1, key="cc_months")

    cc_m = FREQ_OPTIONS[cc_freq_label]
    cc_ear = ear_from_apr(cc_apr / 100, cc_m)
    cc_years = cc_months / 12
    cc_fv = fv_single(cc_balance, cc_ear, cc_years)

    st.divider()
    m1, m2, m3 = st.columns(3)
    m1.metric("Stated APR", fmt_pct(cc_apr / 100))
    m2.metric("True Effective Annual Rate (EAR)", fmt_pct(cc_ear))
    m3.metric(f"Balance after {cc_months} months (unpaid)", fmt_money_plain(cc_fv))

    st.success(f"A card advertised at **{cc_apr:.2f}% APR** compounded {cc_freq_label.lower()} actually costs "
               f"**{fmt_pct(cc_ear)}** per year. Carrying {fmt_money(cc_balance)} with no payments for "
               f"{cc_months} months would grow it to **{fmt_money(cc_fv)}** — "
               f"{fmt_money(cc_fv - cc_balance)} in interest alone.")
    st.caption("By contrast, paying the statement balance in full each month typically avoids interest entirely "
               "on new purchases — the gap above is the cost of carrying a balance, not of using a card responsibly.")

    with st.expander("📊 See this APR's EAR at every compounding frequency"):
        comp_rows_cc = [{"Compounding": lbl, "EAR": ear_from_apr(cc_apr / 100, freq)} for lbl, freq in FREQ_OPTIONS.items()]
        comp_df_cc = pd.DataFrame(comp_rows_cc)
        st.dataframe(comp_df_cc.assign(EAR=comp_df_cc["EAR"].apply(fmt_pct)), use_container_width=True, hide_index=True)

    summary_rows_cc = [
        ("Stated APR", fmt_pct(cc_apr / 100)), ("Compounding Frequency", cc_freq_label),
        ("True EAR", fmt_pct(cc_ear)), ("Starting Balance", fmt_money_plain(cc_balance)),
        ("Months Projected", cc_months), (f"Balance After {cc_months} Months", fmt_money_plain(cc_fv)),
        ("Interest Accrued", fmt_money_plain(cc_fv - cc_balance)),
    ]
    pdf_buf_cc = build_simple_pdf("Credit Card: APR vs. EAR", "Worked example", summary_rows_cc,
                                   footnote="Generated by the Time Value of Money Calculator.")
    st.download_button("⬇️ PDF Report", data=pdf_buf_cc, file_name="credit_card_apr_ear.pdf", mime="application/pdf")
